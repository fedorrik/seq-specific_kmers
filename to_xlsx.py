import os
import pandas as pd
from sys import argv

directory = argv[1]
xlsx_file_path = argv[2]

# Get list of .cnt files in the specified directory
tsv_files = [file for file in os.listdir(directory) if file.endswith('.cnt')]

# Sort files by the last part of the sheet name (sheet_name.split('_')[-1])
# sheet_name is os.path.splitext(file)[0][:-6]
def sheet_sort_key(file):
    base = os.path.splitext(file)[0][:-6]
    last_part = base.split('_')[-1]
    try:
        return last_part
    except ValueError:
        return last_part  # fallback to string comparison if not numeric

tsv_files = sorted(tsv_files, key=sheet_sort_key)

# Create a new ExcelWriter object
with pd.ExcelWriter(xlsx_file_path, engine='xlsxwriter') as writer:
    # Iterate over all .cnt files
    for file in tsv_files:
        # Read data from each .cnt file into a pandas DataFrame
        df = pd.read_csv(os.path.join(directory, file), sep='\t')
        # Use filename (without extension and trailing pattern) as sheet name
        sheet_name = os.path.splitext(file)[0][:-6]
        # Write data to Excel on a separate sheet
        df.to_excel(writer, sheet_name=sheet_name, index=False)

print("xlsx created")

from openpyxl import load_workbook

# Load the existing Excel file
workbook = load_workbook(xlsx_file_path)

# Iterate over all sheets in the workbook
for sheet in workbook.sheetnames:
    ws = workbook[sheet]
    # Set width of the first column to 30
    ws.column_dimensions['A'].width = 30

# Save changes to the Excel file
workbook.save(xlsx_file_path)

print("first column width set to 30")
